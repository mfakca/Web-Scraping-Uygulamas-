import pandas as pd
import numpy as np
import time 
# Çektiğimiz Verileri getiriyorum.
df = pd.read_excel('data3.xlsx',index_col='Unnamed: 0')


# Kontrol ediyorum.
#print(df)
# print(df.columns)


# Ortalama puanı hesaplıyorum.
df ['ortalamaPuani'] = np.round_(np.sum([df['hizPuani'],df['servisPuani'],df['lezzetPuani']],axis=0)/3 , 2)
# print(df)

df['yorumlar'] = df['yorumlar'].apply(lambda x: x.replace('.',' '))

# Normalization işlemi
df['yorumlar'] = df['yorumlar'].apply(lambda x: x.lower())
print(df.yorumlar)


# Tokenize işlemi
'''
from zemberek import TurkishTokenizer


tokenizer = TurkishTokenizer.DEFAULT

deneme = tokenizer.tokenize(df['yorumlar'][1])
for i in deneme:
    print(i.content)
    
'''



# Stop-words'lerin kaldırılması.

import nltk

#nltk.download('stopwords')

from nltk.corpus import stopwords
stopWords = set(stopwords.words('turkish'))


df['yorumlar'] = df['yorumlar'].apply(lambda x: ' '.join([i for i in x.split() if i not in stopWords]))

#print(df.yorumlar)



# Noktalama işaretlerinin kaldırılması.
tokenizer = nltk.RegexpTokenizer(r"\w+")

df['yorumlar'] = df['yorumlar'].apply(lambda x: ' '.join(tokenizer.tokenize(x)))


#print(df.yorumlar)





# Sayıların kaldırılması.
from string import digits 

remove_digits = str.maketrans('', '', digits) 



df['yorumlar'] = df['yorumlar'].apply(lambda x: x.translate(remove_digits))

#print(df.yorumlar).



# Silinen karakterlerden dolayı kalan çift boşlukları tek boşluğa çeviriyorum.
df['yorumlar'] = df['yorumlar'].apply(lambda x: x.replace('  ',' '))
#print(df.yorumlar)









# Kelime köklerini alma
import jpype
# JVM başlat
# Aşağıdaki adresleri java sürümünüze ve jar dosyasının bulunduğu klasöre göre değiştirin
jpype.startJVM("C:/Program Files/Java/jdk1.8.0_221/jre/bin/server/jvm.dll",
         "-Djava.class.path=C:/Users/MehmetFatihAKCA/Desktop/projeler/Web Scrapping - yemek/zemberek-tum-2.0.jar", "-ea")
# Türkiye Türkçesine göre çözümlemek için gerekli sınıfı hazırla
Tr = jpype.JClass("net.zemberek.tr.yapi.TurkiyeTurkcesi")
# tr nesnesini oluştur
tr = Tr()
# Zemberek sınıfını yükle
Zemberek = jpype.JClass("net.zemberek.erisim.Zemberek")
# zemberek nesnesini oluştur
zemberek = Zemberek(tr)
#Çözümlenecek örnek kelimeleri belirle


for cou,kelimeler in enumerate(df.yorumlar):
    yorumKok=[]
    for kelime in kelimeler.split():
        if kelime.strip()>'':
            yanit = zemberek.kelimeCozumle(kelime)
            
            if yanit:
                print("{}".format(yanit[0].kok()).split()[0])
                yorumKok.append(f"{str((yanit[0].kok())).split()[0]}")
                
            else:
                print("{} ÇÖZÜMLENEMEDİ".format(kelime))
                yorumKok.append(kelime)
    df['yorumlar'][cou]=' '.join(yorumKok)
   
#print(df['yorumlar'].apply(lambda x: ''.join([zemberek.kelimeCozumle(kelime)[0] for kelime in x.split() if kelime.strip()>'' ])))
    
#JVM kapat



#print(df.yorumlar)







# Yorumlar içerisindeki geçen en çok 10 kelimeyi 100 ile ölçekledim.
from collections import Counter

all_freq={}
for cou,x in enumerate(df.yorumlar):
    yorumlar_10=""
    for i in x.split(): 
        if i in all_freq: 
            all_freq[i] += 1
        else: 
            all_freq[i] = 1
    print ("Count of all characters in GeeksforGeeks is :\n "
                                        +  str(all_freq)) 
    
    
    c = Counter(all_freq)
    print(c.most_common()[:10])

    
    c = Counter(all_freq)
    for i in c.most_common()[:10]:
        yorumlar_10+=(i[0] + " " ) * i[1] 
    df.yorumlar[cou]=yorumlar_10


import random   
for num,i in enumerate(df.yorumlar):
    try:
        df.yorumlar[num] = " ".join(random.sample(i.split(),100))
    except: 
        df.drop(axis=0,index=num,inplace=True)
        
        print("XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX\n",i.split())








# Bag of Words
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer

 
 
CountVec = CountVectorizer(ngram_range=(1,1))
#transform
Count_data = CountVec.fit_transform(df.yorumlar)
 
#create dataframe
cv_dataframe=pd.DataFrame(Count_data.toarray(),columns=CountVec.get_feature_names())


# Yorumları sayıya çevirdiğimiz için artık buna gerek yok.
df.drop(['yorumlar'],axis=1,inplace=True)


# Asıl dataframe ile Bag of Words'u birleştiriyorum.
data = pd.concat([df,cv_dataframe],axis=1)


# nan değer varsa atıyorum.
data.dropna(axis=0,inplace=True)




data["ortalamaPuani"] = [round(i,1) for i in data["ortalamaPuani"]]
data.to_excel('son.xlsx')

import openpyxl
from openpyxl.styles import PatternFill
wb = openpyxl.load_workbook('son.xlsx')
sheet = wb.active


for i in range(len(data['hizPuani'])):
    sheet[f'H{i+2}'].fill = PatternFill(start_color="FFFF00", fill_type = "solid")


wb.save('son.xlsx')






